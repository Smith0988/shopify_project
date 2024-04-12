import os
import re
import shutil
import sys

import openpyxl
from openpyxl.styles import Alignment


def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)


temp_excel_form = resource_path("temp_data\\temp_form.xlsx")


def write_to_excel(data):
    excel_file = "temp_form.xlsx"
    if not os.path.exists(excel_file):
        source_excel = resource_path("temp_data\\temp_form.xlsx")
        shutil.copy(source_excel, excel_file)
    workbook = openpyxl.load_workbook(excel_file, data_only=True)
    sheet = workbook["Sheet1"]
    next_row = sheet.max_row + 1

    col_1_data = data[0]
    col_2_data = data[1]
    col_3_data = data[2]
    col_4_data = data[3]
    sheet.cell(row=next_row, column=1, value=col_1_data).alignment = Alignment(horizontal='center')
    sheet.cell(row=next_row, column=2, value=col_2_data).alignment = Alignment(horizontal='center')
    sheet.cell(row=next_row, column=3, value=col_3_data).alignment = Alignment(horizontal='center')
    sheet.cell(row=next_row, column=4, value=col_4_data).alignment = Alignment(horizontal='center')

    # Lưu tệp Excel
    workbook.save(excel_file)




def conver_link_to_list(sku, link1, link2, link3, text):

    l1 = ['6pcs5in-i-dont-care-what-the-bible-says-bumper-sticker-vinyl-decal-waterproof-stickers-for-car-truck-vehicl-laptop', "6Pcs(5IN) I Don't Care What The Bible Says Bumper Sticker", '6Pcs Vinyl decal stickers (5(in) x 1.67(in) x 0.03(in) each pack)', 'Chansenchine', 'Vehicles & Parts > Vehicle Parts & Accessories > Vehicle Maintenance, Care & Decor > Vehicle Decor > Bumper Stickers', ' ', ' ', True, 'Title', 'Default Title', ' ', ' ', ' ', ' ', 'SKU20240409_141952', '0.0', 'shopify', 50, 'continue', 'manual', '13.00', '16.00', True, True, ' ', 'https://cdn.shopify.com/s/files/1/0618/9655/6588/files/SKU20240409_141952_IDONTCAREWHATTHEBIBLESAYScopy_3d58742c-87a3-411a-af12-660e66ac29bc.jpg?v=1712648695', 1, ' ', False, ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', 'kg', ' ', ' ', True, ' ', ' ', True, ' ', ' ', True, ' ', ' ', 'active']
    l2 = ['6pcs5in-i-dont-care-what-the-bible-says-bumper-sticker-vinyl-decal-waterproof-stickers-for-car-truck-vehicl-laptop', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', 'https://cdn.shopify.com/s/files/1/0618/9655/6588/files/SKU20240409_141952_IDONTCAREWHATTHEBIBLESAYScopy.jpg?v=1712648694', 2, ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ']
    l3 = ['6pcs5in-i-dont-care-what-the-bible-says-bumper-sticker-vinyl-decal-waterproof-stickers-for-car-truck-vehicl-laptop', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', 'https://cdn.shopify.com/s/files/1/0618/9655/6588/files/SKU20240409_141952_IDONTCAREWHATTHEBIBLESAYScopy_20e3bd3c-8973-43ec-8694-d345d8b1d693.jpg?v=1712648694', 3, ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ']

    text_anpha = re.sub(r'[^a-zA-Z0-9\s]', '', text)
    # Thay thế khoảng trắng bằng dấu -
    text_no_space = text_anpha.replace(' ', '-')
    # Chuyển đổi thành chữ thường
    text_handle = text_no_space.lower()

    l1[0] = text_handle
    l2[0] = text_handle
    l3[0] = text_handle

    l1[1] = text

    l1[14] = sku

    l1[25] = link1
    l2[25] = link2
    l3[25] = link3

    write_list_to_excel(l1, l2, l3)






def write_list_to_excel(list1, list2, list3):

    ten_file = "temp_form.xlsx"
    if not os.path.exists(ten_file):
        source_excel = resource_path("temp_data\\temp_form.xlsx")
        shutil.copy(source_excel, ten_file)

    try:
        # Mở workbook đã có nếu tồn tại, nếu không tạo mới
        wb = openpyxl.load_workbook(ten_file)
    except FileNotFoundError:
        wb = openpyxl.Workbook()

    # Chọn sheet đầu tiên hoặc tạo mới nếu chưa có
    sheet = wb.active

    # Tìm hàng cuối cùng có dữ liệu
    last_row = sheet.max_row + 1

    # Ghi list1 vào hàng tiếp theo
    for col, value in enumerate(list1, start=1):
        sheet.cell(row=last_row, column=col, value=value)

    # Ghi list2 vào hàng tiếp theo
    last_row += 1
    for col, value in enumerate(list2, start=1):
        sheet.cell(row=last_row, column=col, value=value)

    # Ghi list3 vào hàng tiếp theo
    last_row += 1
    for col, value in enumerate(list3, start=1):
        sheet.cell(row=last_row, column=col, value=value)

    # Lưu workbook vào tệp Excel
    wb.save(ten_file)




if __name__ == "__main__":
    # Sử dụng hàm để ghi dữ liệu vào tệp Excel

    l1 = ['6pcs5in-i-dont-care-what-the-bible-says-bumper-sticker-vinyl-decal-waterproof-stickers-for-car-truck-vehicl-laptop', "6Pcs(5IN) I Don't Care What The Bible Says Bumper Sticker", '6Pcs Vinyl decal stickers (5(in) x 1.67(in) x 0.03(in) each pack)', 'Chansenchine', 'Vehicles & Parts > Vehicle Parts & Accessories > Vehicle Maintenance, Care & Decor > Vehicle Decor > Bumper Stickers', ' ', ' ', True, 'Title', 'Default Title', ' ', ' ', ' ', ' ', 'SKU20240409_141952', '0.0', 'shopify', 50, 'continue', 'manual', '13.00', '16.00', True, True, ' ', 'https://cdn.shopify.com/s/files/1/0618/9655/6588/files/SKU20240409_141952_IDONTCAREWHATTHEBIBLESAYScopy_3d58742c-87a3-411a-af12-660e66ac29bc.jpg?v=1712648695', 1, ' ', False, ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', 'kg', ' ', ' ', True, ' ', ' ', True, ' ', ' ', True, ' ', ' ', 'active']
    l2 = ['6pcs5in-i-dont-care-what-the-bible-says-bumper-sticker-vinyl-decal-waterproof-stickers-for-car-truck-vehicl-laptop', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', 'https://cdn.shopify.com/s/files/1/0618/9655/6588/files/SKU20240409_141952_IDONTCAREWHATTHEBIBLESAYScopy.jpg?v=1712648694', 2, ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ']
    l3 = ['6pcs5in-i-dont-care-what-the-bible-says-bumper-sticker-vinyl-decal-waterproof-stickers-for-car-truck-vehicl-laptop', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', 'https://cdn.shopify.com/s/files/1/0618/9655/6588/files/SKU20240409_141952_IDONTCAREWHATTHEBIBLESAYScopy_20e3bd3c-8973-43ec-8694-d345d8b1d693.jpg?v=1712648694', 3, ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ']

    #write_list_to_excel(l1,l2,l3)